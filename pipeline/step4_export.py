"""
Step 4: Clean up and export the final lead list.
Outputs a styled Excel file with metadata banner, Excel Table, hyperlinks,
number formatting, and level-based row coloring.
"""

import logging
from datetime import datetime
from pathlib import Path

import pandas as pd
import requests
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")

_OUTPUTS    = Path(__file__).parent.parent / "Outputs"
INPUT_PATH  = _OUTPUTS / "with_contacts.csv"
OUTPUT_PATH = _OUTPUTS / "final_leads.xlsx"

COLUMN_ORDER = [
    "contact_name", "contact_title", "contact_level",
    "email", "linkedin",
    "company", "company_website", "company_address",
    "company_phone", "company_reviews_avg", "company_reviews_count",
    "exported_at",
]

COLUMN_LABELS = {
    "contact_name":          "Contact Name",
    "contact_title":         "Title",
    "contact_level":         "Level",
    "email":                 "Email",
    "linkedin":              "LinkedIn",
    "company":               "Company",
    "company_website":       "Website",
    "company_address":       "Address",
    "company_phone":         "Phone",
    "company_reviews_avg":   "Avg Rating",
    "company_reviews_count": "Review Count",
    "exported_at":           "Exported At",
}

LEVEL_SORT    = {"level1": 0, "level2": 1}
LEVEL_DISPLAY = {"level1": "Senior (L1)", "level2": "Mid (L2)"}

META_COLOR   = "0D2137"   # dark navy — metadata banner
HEADER_COLOR = "1F4E79"   # dark blue — column headers
LEVEL1_COLOR = "D6E4F0"   # light blue — Senior (L1) rows
LEVEL2_COLOR = "FFFFFF"   # white      — Mid (L2) rows
LINK_COLOR   = "1155CC"   # hyperlink blue


def _col_width(series, header):
    max_data = series.astype(str).str.len().max() if len(series) else 0
    return min(max(max_data, len(header)) + 4, 60)


def run(config):
    df = pd.read_csv(INPUT_PATH)

    # Drop placeholder rows
    df = df[df["contact_name"].notna() & (df["contact_name"].astype(str).str.strip() != "")].copy()

    # Sort using original level values before display transform
    df["_sort"] = df["contact_level"].map(LEVEL_SORT).fillna(2)
    df = df.sort_values(["company", "_sort"]).drop(columns=["_sort"])
    df = df.reset_index(drop=True)

    # Preserve original df for backend push (original level values, no exported_at)
    df_backend = df.copy()

    # Apply display transformations for Excel only
    df["contact_level"] = df["contact_level"].map(LEVEL_DISPLAY).fillna(df["contact_level"])
    exported_at_str = datetime.now().strftime("%d %b %Y, %I:%M %p")
    df["exported_at"] = exported_at_str

    # Keep only known columns in order
    cols = [c for c in COLUMN_ORDER if c in df.columns]
    df   = df[cols].copy()

    n_rows   = len(df)
    n_cols   = len(cols)
    last_col = get_column_letter(n_cols)

    # ── Write Excel ──────────────────────────────────────────────────────────
    with pd.ExcelWriter(OUTPUT_PATH, engine="openpyxl") as writer:
        # startrow=1 → header lands on row 2, data starts at row 3
        df.rename(columns=COLUMN_LABELS).to_excel(
            writer, index=False, sheet_name="Leads", startrow=1
        )
        ws = writer.sheets["Leads"]

        thin_border = Border(
            bottom=Side(style="thin", color="BBBBBB"),
            right=Side(style="thin",  color="BBBBBB"),
        )

        # ── Metadata banner (row 1) ──────────────────────────────────────────
        ws.merge_cells(f"A1:{last_col}1")
        meta           = ws["A1"]
        meta.value     = (
            f"SortMyPrep Leads Export   |   {n_rows} leads   |   "
            f"Generated: {exported_at_str}"
        )
        meta.font      = Font(bold=True, color="FFFFFF", size=12)
        meta.fill      = PatternFill("solid", fgColor=META_COLOR)
        meta.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 26

        # ── Column header row (row 2) ────────────────────────────────────────
        for cell in ws[2]:
            cell.font      = Font(bold=True, color="FFFFFF", size=11)
            cell.fill      = PatternFill("solid", fgColor=HEADER_COLOR)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = thin_border
        ws.row_dimensions[2].height = 22

        # ── Column index helpers ─────────────────────────────────────────────
        def col_idx(key):
            return cols.index(key) + 1 if key in cols else None

        level_col    = col_idx("contact_level")
        linkedin_col = col_idx("linkedin")
        website_col  = col_idx("company_website")
        email_col    = col_idx("email")
        rating_col   = col_idx("company_reviews_avg")
        count_col    = col_idx("company_reviews_count")

        # ── Data rows (row 3 onwards) ────────────────────────────────────────
        l1_fill    = PatternFill("solid", fgColor=LEVEL1_COLOR)
        l2_fill    = PatternFill("solid", fgColor=LEVEL2_COLOR)
        data_align = Alignment(vertical="center", wrap_text=False)

        for row_idx in range(3, n_rows + 3):
            level_val = ws.cell(row=row_idx, column=level_col).value if level_col else ""
            row_fill  = l1_fill if level_val == "Senior (L1)" else l2_fill

            for cell in ws[row_idx]:
                cell.fill      = row_fill
                cell.alignment = data_align
                cell.border    = thin_border
            ws.row_dimensions[row_idx].height = 18

            # Clickable hyperlinks — LinkedIn and Website
            for ci in filter(None, [linkedin_col, website_col]):
                c = ws.cell(row=row_idx, column=ci)
                if c.value and str(c.value).startswith("http"):
                    c.hyperlink = str(c.value)
                    c.font = Font(color=LINK_COLOR, underline="single")

            # Clickable mailto — Email
            if email_col:
                c = ws.cell(row=row_idx, column=email_col)
                if c.value and "@" in str(c.value):
                    c.hyperlink = f"mailto:{c.value}"
                    c.font = Font(color=LINK_COLOR, underline="single")

            # Number formats
            if rating_col:
                ws.cell(row=row_idx, column=rating_col).number_format = "0.0"
            if count_col:
                ws.cell(row=row_idx, column=count_col).number_format = "#,##0"

        # ── Excel Table (filter dropdowns on every column) ───────────────────
        tab = Table(displayName="LeadsTable", ref=f"A2:{last_col}{n_rows + 2}")
        tab.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=False,
            showColumnStripes=False,
        )
        ws.add_table(tab)

        # ── Column widths ────────────────────────────────────────────────────
        for i, key in enumerate(cols, start=1):
            label = COLUMN_LABELS.get(key, key)
            ws.column_dimensions[get_column_letter(i)].width = _col_width(df[key], label)

        # Freeze below metadata banner + header
        ws.freeze_panes = "A3"

    logging.info(f"Step 4 complete. {n_rows} leads exported → {OUTPUT_PATH}")

    backend_url = config.get("backend_url", "").strip()
    if backend_url:
        _push_to_backend(df_backend, backend_url)


def _push_to_backend(df: pd.DataFrame, backend_url: str):
    import json
    leads = json.loads(df.to_json(orient="records"))
    try:
        resp = requests.post(
            f"{backend_url.rstrip('/')}/api/leads/bulk",
            json={"leads": leads},
            timeout=30,
        )
        resp.raise_for_status()
        result = resp.json()
        logging.info(
            f"Backend sync: {result.get('inserted', 0)} new, "
            f"{result.get('skipped', 0)} already existed"
        )
    except Exception as e:
        logging.warning(f"Backend sync failed (leads still saved locally): {e}")
